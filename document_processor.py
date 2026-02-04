"""
Document processing module for extracting text from various file formats
"""
import os
from pathlib import Path
from typing import List, Dict, Optional
import logging

# PDF processing
try:
    import PyPDF2
except ImportError:
    PyPDF2 = None

# Word document processing
try:
    from docx import Document
except ImportError:
    Document = None

# Excel processing
try:
    import openpyxl
except ImportError:
    openpyxl = None

logger = logging.getLogger(__name__)


class DocumentProcessor:
    """Processes various document formats and extracts text"""
    
    def __init__(self):
        self.supported_formats = {
            '.pdf': self._process_pdf,
            '.docx': self._process_docx,
            '.doc': self._process_docx,  # Treat .doc as .docx (may need conversion)
            '.txt': self._process_txt,
            '.xlsx': self._process_excel,
            '.xls': self._process_excel,
        }
    
    def process_file(self, file_path: str) -> Dict[str, any]:
        """
        Process a file and extract its content
        
        Args:
            file_path: Path to the file
            
        Returns:
            Dictionary with 'text', 'metadata', and 'success' fields
        """
        file_path = Path(file_path)
        if not file_path.exists():
            return {
                'success': False,
                'error': f'File not found: {file_path}',
                'text': '',
                'metadata': {}
            }
        
        extension = file_path.suffix.lower()
        
        if extension not in self.supported_formats:
            return {
                'success': False,
                'error': f'Unsupported file format: {extension}',
                'text': '',
                'metadata': {}
            }
        
        try:
            processor = self.supported_formats[extension]
            text = processor(str(file_path))
            
            metadata = {
                'filename': file_path.name,
                'file_path': str(file_path),
                'file_size': file_path.stat().st_size,
                'file_type': extension,
            }
            
            return {
                'success': True,
                'text': text,
                'metadata': metadata,
                'error': None
            }
        except Exception as e:
            logger.error(f"Error processing file {file_path}: {str(e)}")
            return {
                'success': False,
                'error': f'Error processing file: {str(e)}',
                'text': '',
                'metadata': {}
            }
    
    def _process_pdf(self, file_path: str) -> str:
        """Extract text from PDF file"""
        if PyPDF2 is None:
            raise ImportError("PyPDF2 is not installed")
        
        text = ""
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
        return text.strip()
    
    def _process_docx(self, file_path: str) -> str:
        """Extract text from Word document"""
        if Document is None:
            raise ImportError("python-docx is not installed")
        
        doc = Document(file_path)
        text = []
        for paragraph in doc.paragraphs:
            text.append(paragraph.text)
        return "\n".join(text).strip()
    
    def _process_txt(self, file_path: str) -> str:
        """Extract text from plain text file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read()
        except UnicodeDecodeError:
            # Try with different encoding
            with open(file_path, 'r', encoding='latin-1') as file:
                return file.read()
    
    def _process_excel(self, file_path: str) -> str:
        """Extract text from Excel file"""
        if openpyxl is None:
            raise ImportError("openpyxl is not installed")
        
        workbook = openpyxl.load_workbook(file_path)
        text_parts = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_parts.append(f"Sheet: {sheet_name}\n")
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text_parts.append(row_text)
            text_parts.append("\n")
        
        return "\n".join(text_parts).strip()
